import requests
from bs4 import BeautifulSoup
import openpyxl
from urllib.parse import urljoin

def clean_text(text):
    return ''.join(char for char in text if char.isprintable())

url = 'https://wistev.de/'

# Set a user-agent to mimic a web browser
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

try:
    response = requests.get(url, headers=headers, allow_redirects=True)
    response.raise_for_status()

    # Manually follow redirects up to a limit
    max_redirects = 10
    current_redirects = 0
    while response.is_redirect and current_redirects < max_redirects:
        absolute_url = urljoin(url, response.headers['Location'])
        response = requests.get(absolute_url, headers=headers, allow_redirects=True)
        response.raise_for_status()
        current_redirects += 1

    # Parse the HTML content
    soup = BeautifulSoup(response.text, 'html.parser')

    # Extract links
    links = [link.get('href') for link in soup.find_all("a")]

    # Create a new Excel workbook and sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Links with Text"

    # Add headers
    sheet['A1'] = 'Source URL'
    sheet['B1'] = 'Text Content'

    # Write links and text to Excel sheet
    for index, link in enumerate(links, start=2):
        absolute_url = urljoin(url, link)
        
        try:
            link_response = requests.get(absolute_url, headers=headers)
            link_response.raise_for_status()
            
            link_soup = BeautifulSoup(link_response.text, 'html.parser')
            
            text_content = clean_text(link_soup.get_text(strip=True))
            
            sheet.cell(row=index, column=1, value=absolute_url)
            sheet.cell(row=index, column=2, value=text_content)
        
        except requests.exceptions.RequestException as e:
            print(f"Error accessing {absolute_url}: {e}")

    # Save the workbook to a file
    wb.save("links_with_text.xlsx")

    print(f"Links and text written to links_with_text.xlsx successfully.")

except requests.exceptions.RequestException as e:
    print(f"Error: {e}")
